from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.contrib.auth import authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.http import HttpResponse
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Question, ExamSession, Answer, DEPARTMENT_CHOICES
from .forms import MockRegistrationForm, FinalRegistrationForm, EnglishRegistrationForm

# 1. ASOSIY SAHIFA
def index_view(request):
    """3 ta bo'limli asosiy dashboard sahifasi"""
    return render(request, 'exam/index.html')

# 2. LOGISTICS MOCK REGISTRATION
def register_mock_view(request):
    """Logistika Mock Exam uchun ro'yxatdan o'tish[cite: 2]"""
    if request.method == 'POST':
        form = MockRegistrationForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.exam_type = 'mock'
            session.exam_date = timezone.now().date()
            session.save()
            return redirect('exam_start', session_id=session.id)
    else:
        form = MockRegistrationForm()
    return render(request, 'exam/register.html', {
        'form': form, 
        'title': 'Practice Session', 
        'subtitle': 'Mock Registration'
    })

# 3. ENGLISH SESSION REGISTRATION (YANGI)
def register_english_view(request):
    """Ingliz tili sessiyasi uchun alohida ro'yxatdan o'tish[cite: 3]"""
    if request.method == 'POST':
        form = EnglishRegistrationForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.exam_type = 'mock' # English sessionlar mock sifatida saqlanadi
            session.exam_date = timezone.now().date()
            # ChoiceField bitta string qiymat qaytaradi
            session.save()
            return redirect('exam_start', session_id=session.id)
    else:
        form = EnglishRegistrationForm()
    
    return render(request, 'exam/register_eng.html', {
        'form': form,
        'title': 'English Session',
        'subtitle': 'Registration & Module Selection'
    })

# 4. FINAL EXAM KODINI TEKSHIRISH
def final_access_view(request):
    """Final Examga kirish uchun o'qituvchi loginini tekshirish[cite: 2]"""
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None and user.is_staff:
            request.session['final_allowed'] = True
            return redirect('register_final')
        else:
            messages.error(request, "O'qituvchi login yoki paroli xato!")
    return render(request, 'exam/final_access.html')

# 5. FINAL REGISTRATION
def register_final_view(request):
    """Logistika Final Exam uchun ro'yxatdan o'tish[cite: 2]"""
    if not request.session.get('final_allowed'):
        return redirect('final_access')
    if request.method == 'POST':
        form = FinalRegistrationForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.exam_type = 'final'
            session.save()
            del request.session['final_allowed']
            return redirect('exam_start', session_id=session.id)
    else:
        form = FinalRegistrationForm()
    return render(request, 'exam/register.html', {
        'form': form, 
        'title': 'Final Exam', 
        'subtitle': 'Official Certification'
    })

# 6. IMTIHON JARAYONI (6+4+2 VA 10+5+5)
@ensure_csrf_cookie
def exam_view(request, session_id):
    """Savollarni tasodifiy saralash va imtihon jarayoni[cite: 2]"""
    session = get_object_or_404(ExamSession, id=session_id, is_complete=False)
    
    if request.method == 'POST':
        question_ids = request.POST.getlist('question_ids')
        for q_id in question_ids:
            field_key = f"q_{q_id}"
            value = request.POST.get(field_key, "").strip()
            if value:
                q = Question.objects.get(id=q_id)
                ans, _ = Answer.objects.get_or_create(session=session, question=q)
                if q.question_type == 'mcq':
                    ans.selected_option = value
                else:
                    ans.answer_text = value
                ans.save()
        session.submitted_at = timezone.now()
        session.is_complete = True
        session.save()
        return redirect('exam_done')

    selected_depts = session.get_departments_list()
    is_english = any(d.startswith('eng_') for d in selected_depts)
    
    all_qs = []

    if is_english:
        # INGLIZ TILI LOGIKASI (10+5+5)[cite: 3]
        for d in selected_depts:
            if d.startswith('eng_'):
                mcqs = list(Question.objects.filter(department=d, question_type='mcq', is_active=True))
                all_qs.extend(random.sample(mcqs, min(10, len(mcqs))))
                
                opens = list(Question.objects.filter(department=d, question_type='open', is_active=True))
                all_qs.extend(random.sample(opens, min(5, len(opens))))
                
                scrambles = list(Question.objects.filter(department=d, question_type='scramble', is_active=True))
                all_qs.extend(random.sample(scrambles, min(5, len(scrambles))))
        
        template_name = 'exam/exam_eng.html'
        exam_time = 30 * 60
    else:
        # LOGISTIKA LOGIKASI (6+4+2)[cite: 2]
        mcq_qs = []
        if selected_depts:
            per_dept = 6 // len(selected_depts)
            for d in selected_depts:
                qs = list(Question.objects.filter(department=d, question_type='mcq', is_active=True))
                if qs: mcq_qs.extend(random.sample(qs, min(per_dept, len(qs))))
        
        if len(mcq_qs) < 6:
            rem = 6 - len(mcq_qs)
            extra = list(Question.objects.filter(department__in=selected_depts, question_type='mcq', is_active=True).exclude(id__in=[q.id for q in mcq_qs]))
            if extra: mcq_qs.extend(random.sample(extra, min(rem, len(extra))))

        open_pool = list(Question.objects.filter(department__in=selected_depts, question_type='open', is_active=True))
        open_qs = random.sample(open_pool, min(4, len(open_pool))) if open_pool else []
        
        career_pool = list(Question.objects.filter(department='career', is_active=True))
        career_qs = random.sample(career_pool, min(2, len(career_pool))) if career_pool else []
        
        all_qs = mcq_qs + open_qs + career_qs
        template_name = 'exam/exam.html'
        exam_time = 20 * 60

    all_qs.sort(key=lambda q: q.order)

    return render(request, template_name, {
        'session': session,
        'questions': all_qs,
        'exam_duration': exam_time,
        'warning_time': 5 * 60,
    })

# 7. TEACHER'S PANEL (ADMIN RESULTS)
@login_required(login_url='/admin/login/')
def admin_results_view(request):
    """Natijalarni ko'rsatish[cite: 2]"""
    if not request.user.is_staff:
        return redirect('home')

    sessions = ExamSession.objects.prefetch_related('answers__question').all().order_by('-submitted_at')
    dept_map = dict(DEPARTMENT_CHOICES)
    
    rows = []
    for session in sessions:
        ans_data = []
        for ans in session.answers.select_related('question').all().order_by('question__order'):
            val = ans.selected_option.upper() if ans.selected_option else ans.answer_text
            ans_data.append({
                'dept': ans.question.get_department_display(),
                'title': ans.question.title,
                'question_text': ans.question.text,
                'value': val,
            })
        rows.append({
            'session': session,
            'dept_display': ', '.join([dept_map.get(d, d) for d in session.get_departments_list()]),
            'answered': ans_data,
        })
    
    response = render(request, 'exam/admin_results.html', {'rows': rows})
    logout(request) 
    return response

# 8. EXCEL EXPORT
@staff_member_required
def export_excel_view(request):
    """Natijalarni Excelga yuklash[cite: 2]"""
    sessions = ExamSession.objects.prefetch_related('answers__question').all()
    dept_map = dict(DEPARTMENT_CHOICES)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed MBA Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3c5e")
    top_align = Alignment(vertical="top", wrap_text=True)

    headers = ["Type", "Full Name", "Phone", "Date", "Group", "Teacher", "Departments"]
    for i in range(1, 25): headers.append(f"Q&A {i}") # Ustunlarni 24 taga kengaytirdik

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill = header_font, header_fill
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20 if col <= 7 else 60

    for row_num, session in enumerate(sessions, 2):
        ws.cell(row=row_num, column=1, value=session.get_exam_type_display()).alignment = top_align
        ws.cell(row=row_num, column=2, value=session.full_name).alignment = top_align
        ws.cell(row=row_num, column=3, value=session.phone).alignment = top_align
        ws.cell(row=row_num, column=4, value=str(session.exam_date)).alignment = top_align
        ws.cell(row=row_num, column=5, value=session.group).alignment = top_align
        ws.cell(row=row_num, column=6, value=session.teacher).alignment = top_align
        ws.cell(row=row_num, column=7, value=', '.join([dept_map.get(d, d) for d in session.get_departments_list()])).alignment = top_align

        ans_col = 8
        for ans in session.answers.select_related('question').all().order_by('question__order'):
            if ans_col > 31: break
            std_ans = ans.selected_option.upper() if ans.selected_option else ans.answer_text
            ws.cell(row=row_num, column=ans_col, value=f"Q: {ans.question.text}\n\nA: {std_ans}").alignment = top_align
            ans_col += 1

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="MBA_Results_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

# 9. DONE VIEW
def done_view(request):
    """Imtihon yakunlangandagi sahifa[cite: 2]"""
    return render(request, 'exam/done.html')