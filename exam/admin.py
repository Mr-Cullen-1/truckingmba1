from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.utils import timezone
from django.urls import path
from django.shortcuts import render
from django import forms
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Question, ExamSession, Answer, DEPARTMENT_CHOICES


# Excel faylni admin panel orqali qabul qilish uchun kichik forma
class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Excel faylini tanlang (.xlsx)")


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ['order', 'department', 'question_type', 'title', 'is_active']
    list_display_links = ['title']
    list_editable = ['order', 'is_active']
    list_filter = ['department', 'question_type', 'is_active']
    search_fields = ['title', 'text']
    fieldsets = (
        ('Question Info', {
            'fields': ('department', 'question_type', 'title', 'text', 'order', 'is_active')
        }),
        ('MCQ Options (leave blank for open questions)', {
            'fields': ('option_a', 'option_b', 'option_c', 'option_d'),
            'classes': ('collapse',),
        }),
    )

    # Shaxsiy tugmani chiqarish uchun admin template'ni ulaymiz
    change_list_template = "admin/question_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.import_excel, name='import_excel'),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    
                    count = 0
                    # Sarlavhadan keyingi (2-qator) ma'lumotlarni o'qishni boshlaymiz
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]:
                            continue
                            
                        dept, q_type, title, text, opt_a, opt_b, opt_c, opt_d, order, is_active = row[:10]
                        
                        # Ma'lumotlarni tozalash va standart qiymat berish
                        order = int(order) if order is not None else 0
                        is_active = bool(is_active) if is_active is not None else True
                        
                        # text (savol matni) orqali tekshirib, bor bo'lsa yangilaydi, yo'q bo'lsa yangi ochadi
                        Question.objects.update_or_create(
                            text=str(text).strip(),
                            defaults={
                                'department': str(dept).strip(),
                                'question_type': str(q_type).strip(),
                                'title': str(title).strip(),
                                'option_a': str(opt_a).strip() if opt_a else '',
                                'option_b': str(opt_b).strip() if opt_b else '',
                                'option_c': str(opt_c).strip() if opt_c else '',
                                'option_d': str(opt_d).strip() if opt_d else '',
                                'order': order,
                                'is_active': is_active
                            }
                        )
                        count += 1
                    
                    self.message_user(request, f"Muvaffaqiyatli yakunlandi! {count} ta savol bazaga yuklandi.", messages.SUCCESS)
                    return HttpResponseRedirect("../")
                except Exception as e:
                    self.message_user(request, f"Xatolik yuz berdi: {str(e)}", messages.ERROR)
                    return HttpResponseRedirect(".")
        else:
            form = ExcelImportForm()
        
        context = {
            'form': form,
            'opts': self.model._meta,
            'title': "Exceldan savollarni import qilish"
        }
        return render(request, "admin/excel_import_form.html", context)


class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0
    readonly_fields = ['question', 'selected_option', 'answer_text']
    can_delete = False


def export_sessions_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3c5e")
    center = Alignment(horizontal="center", vertical="center")

    base_headers = ["ID", "Full Name", "Phone", "Date", "Group", "Teacher",
                    "Departments", "Started At", "Submitted At", "Complete"]

    for col, header in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 22

    for row_num, session in enumerate(queryset, 2):
        dept_map = dict(DEPARTMENT_CHOICES)
        dept_display = ', '.join([dept_map.get(d, d) for d in session.get_departments_list()])

        row_data = [
            session.id,
            session.full_name,
            session.phone,
            str(session.exam_date),
            session.group,
            session.teacher,
            dept_display,
            session.started_at.strftime("%Y-%m-%d %H:%M") if session.started_at else "",
            session.submitted_at.strftime("%Y-%m-%d %H:%M") if session.submitted_at else "",
            "Yes" if session.is_complete else "No",
        ]

        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)

        ans_col = len(base_headers) + 1
        for ans in session.answers.select_related('question').all():
            val = ans.selected_option.upper() if ans.selected_option else ans.answer_text
            label = f"[{ans.question.get_department_display()}] {ans.question.title}"
            ws.cell(row=1, column=ans_col, value=label).font = header_font
            ws.cell(row=1, column=ans_col).fill = header_fill
            ws.column_dimensions[ws.cell(row=1, column=ans_col).column_letter].width = 30
            ws.cell(row=row_num, column=ans_col, value=val)
            ans_col += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="exam_results.xlsx"'
    wb.save(response)
    return response

export_sessions_excel.short_description = "📥 Export selected to Excel"


@admin.register(ExamSession)
class ExamSessionAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'group', 'teacher', 'exam_date',
                    'departments', 'started_at', 'submitted_at', 'is_complete']
    list_filter = ['is_complete', 'exam_date', 'group', 'teacher']
    search_fields = ['full_name', 'phone', 'group', 'teacher']
    readonly_fields = ['started_at', 'submitted_at']
    inlines = [AnswerInline]
    actions = [export_sessions_excel]